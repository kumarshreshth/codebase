import openpyxl as op
username=input("enter your username \n")
wb=op.load_workbook('text.xlsx')
wb.active
sheet=wb.sheetnames
ws=wb[sheet[0]]
row1=ws.max_row
count=0
for c in range (1,row1+1):
    val=ws.cell(row=c,column=1)
    if username==val.value:
        count+=1
        print("yes")
        break
if count!=1:
    element=ws.cell(row=row1+1,column=1)
    element.value=username
wb.save("text.xlsx")
#need to integrate password and username module   