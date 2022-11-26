try :
    import openpyxl as op
except :
    pass

def digitValidity(password):
    m = ""
    count = 0
    checklist = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0"]
    for c in password:
        if c in checklist:
            count += 1
            break
    if count == 1:
        return m
    else:
        m = "your password does not contain number"
        return m


def charValidity(password):
    m = ""
    c = 0
    d = 0
    for pas in password:
        if ord(pas) in range(65, 91):
            c += 1
            break
    for pas in password:
        if ord(pas) in range(97, 122):
            d += 1
            break
    if c != 1 and d != 1:
        s = "invalid please enter character"
        m = m + s
    if c != 1 and d == 1:
        s = "please include capital character"
        m = m + s
    if c == 1 and d != 1:
        s = "please include small character"
        m = m + s
    return m


def lengthValidity(password):
    m = ""
    n = len(password)
    if n >= 8:
        return m
    else:
        m = "your password is too short it must contain at least 8 character"
        return m


def specialcharValidity(password):
    m = ""
    count = 0
    checklist = ["!", "@", "#", "$", "%", "^", "&", "*", "(", ")"]
    for c in password:
        if c in checklist:
            count += 1
            break
    if count == 1:
        return m
    else:
        m = "your password does not contain any special character"
        return m


def passwordvalid(password):
    check = []
    digitValid = digitValidity(password)
    charValid = charValidity(password)
    lengthValid = lengthValidity(password)
    specialcharValid = specialcharValidity(password)
    check.extend([digitValid, charValid, lengthValid, specialcharValid])
    n = len(check)
    count = 0
    for c in range(0, n):
        if check[c] == "":
            count += 1
    if count == n:
        print("password is accepted")
        return True
    else:
        print("password is not accepted because :")
        for c in range(0, n):
            if check[c] != "":
                print(check[c])
        return False


def emailformatValidity(email_address):
    m = ""
    count = 0
    checklist = ["@"]
    for c in email_address:
        if c in checklist:
            count += 1
            break
    if count != 1:
        m = "Incorrect Email Address format"
    return m


def emailValidity(email_address):
    check = []
    specialchar = emailformatValidity(email_address)
    check.extend([specialchar])
    n = len(check)
    count = 0
    for c in range(0, n):
        if check[c] == "":
            count += 1
    if count == n:
        print("email address accepted")
        return True
    else:
        print("email address is not accepted due to : ")
        for c in range(0, n):
            if check[c] != "":
                print(check[c])
        return False

try :
    wb = op.load_workbook('test.xlsx')
    wb.active
    sheet = input("enter usertype:\n")
    ws = wb[sheet]
    #check user type
    if sheet == "local" or sheet == "global":
        for c in range(1, ws.max_column + 1):
            element = ws.cell(row=1, column=c)
            if element.value == "PASSWORD":
                k = c
                break
        for c in range(1, ws.max_column + 1):
            element = ws.cell(row=1, column=c)
            if element.value == "USERNAME":
                q = c
                break
        for c in range(1, ws.max_column + 1):
            element = ws.cell(row=1, column=c)
            if element.value == "EMAIL ADDRESS":
                r = c
                break

        #Update/Insert Data
        username = input("enter userId\n")
        password = ""
        count = 0
        row1 = ws.max_row
        for c in range(1, row1 + 1):
            val = ws.cell(row=c, column=q)
            #checking if userId exists
            if username == val.value:
                count += 1
                print("\nUserId exist in " + sheet + " sheet")
                print("enter 1 to update details " +"\nenter 2 to fetch details")
                choice_password = input()
                if choice_password == "2":
                    break
                if choice_password == "1":
                    password = input("enter your password\n")
                    m = passwordvalid(password)
                    element_update_password = ws.cell(row=c, column=k)
                    if m == True:
                        element_update_password.value = password
                    email_address = input("enter your email address\n")
                    j = emailValidity(email_address)
                    element_update_emailaddress = ws.cell(row=c, column=r)
                    if j == True:
                        element_update_emailaddress.value = email_address
                else:
                    print("option not available")
                break
        #userId does not exists
        if count != 1:
            password = input("enter your password\n")
            m = passwordvalid(password)
            if m == True:
                element_username = ws.cell(row=row1 + 1, column=q)
                element_username.value = username
                element_password = ws.cell(row=row1 + 1, column=k)
                element_password.value = password
                email_address = input("enter your email address\n")
                j = emailValidity(email_address)
                element_emailaddress = ws.cell(row=row1 + 1, column=r)
                if j == True:
                    element_emailaddress.value = email_address
        wb.save("test.xlsx")
    else:
        print("Invalid Usertype")
except Exception as e:
    print("Error in execution :" + str(e))


# test scenario : in passwordvalidation what if I enter :,"'{[\|.?/+=-`~or empty space, they are invalid characters, define a function to exclude these, message saying these characters not allowed

# suggestion : define a module called update details where based on count value either update or insert will be done, instead of repeting the code two times
# suggestion :use more appropriate variable names and file names instead on c, j, test, etc


# enhancement 1: add validation in email address that .com should be present at last in email address and there should be characters b/w @ and .com, no special characters in email address except . or - and not grater than 80 characters and no empty space
# enhancement 2: for username no special characters except . or - and not greater than 120 characters
# enhancement 3: while password update if the previous password matches current password, dont update, display message password cant contain old password, not more than 50 characters
# enhancement 4: while password update if password contains username, dont update, saying password cant contain username
# enhancement 5: confirm password module after password validation, and then commit to database


# future development 1: prepare the fetch module function based on user type
# future development 2: prepare mobile number validation and otp module, where a mobile number can be entered/updated only after successfully verified
# future development 3: when password updation only after otp verified provided via mobile number
# future development 4: redesign the whole system removing dependency from user type and search first in local then in global
# future development 5: username does not needs to be unique b/w usertypes, local and global user can have both same username, the id will be different, Gusername for global, Lusername for local, on the same usertype the username has to be unique
# future development 6: remove complete dependency on username as each username will have unique id genrated by random module appended with usertype and 5 letters of username which will be generated and not input
# future development 7: username, password and mobile number all are mandatory fields, if empty delete that entry from database during fetch, displaying correct message
# future development 8: add license expiration date field, based on expiration date, send and email before 20 days to renew system license, send the doc file to email Id, if present else text message on phone
# future development 9: on expired date delete the databse entry and sent email and test message saying contract expired