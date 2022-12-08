try :
    import openpyxl as op
except :
    pass



def usernamecheck(username):
    m=""
    s="it must not contain any special character except - or ."
    n=len(username)
    checklist=[".","-"]
    for c in username:
        if ord(c)==45 or ord(c)==46:
            m=""
            break
        elif ord(c)>=32 and ord(c)<48:
            m=s
            break
        elif ord(c)>=58 and ord(c)<65:
            m=s
            break
        elif ord(c)>=91 and ord(c)<97:
            m=s
            break
        elif ord(c)>=123 and ord(c)<127:
            m=s
            break
        else:
            m=""
    if m=="":
        return True
    else:
        return m





def digitValidity(password):
    m = ""
    count = 0
    checklist = ["1", "2", "3", "4", "5", "6", "7", "8", "9", "0"]
    for c in password:
        if c in checklist:
            count += 1
            break
    if count == 1:
        return m
    else:
        m = "your password does not contain number"
        return m


def charValidity(password):
    m = ""
    c = 0
    d = 0
    for pas in password:
        if ord(pas) in range(65, 91):
            c += 1
            break
    for pas in password:
        if ord(pas) in range(97, 122):
            d += 1
            break
    if c != 1 and d != 1:
        s = "invalid please enter character"
        m = m + s
    if c != 1 and d == 1:
        s = "please include capital character"
        m = m + s
    if c == 1 and d != 1:
        s = "please include small character"
        m = m + s
    return m


def lengthValidity(password):
    m = ""
    n = len(password)
    if n >= 8 and n<50:
        return m
    if n<8:
        m = "your password is too short it must contain at least 8 character"
        return m
    if n>50:
        m="your password is too long it must be less than 50 character"
        return m


def specialcharValidity(password):
    m = ""
    count = 0
    checklist = ["-","`","~","/","_","!", "@", "#", "$", "%", "^", "&",".","<",">","[","]","{","}","|",":",";","'",'"',"*", "(", ")"]
    for c in password:
        if c in checklist:
            count += 1
            break
    if count == 1:
        return m
    else:
        m = "your password does not contain any special character"
        return m



def checkchar(password):
    m=""
    checklist=["?","+","="," "]
    empty=[]
    temp=[]
    for c in password:
        if c in checklist:
            empty.append(c)
    for p in empty:
        if p not in temp:
            temp.append(p)
    for k in temp:
        if k==" ":
            s="your password contain empty space "+" "
        else:
            s="your password contain "+str(k) +" "
        m=m+s
    return m  
    
    
    
def usernameValid(password,username):
    m=""
    list_password=[]
    list_username=[]
    list_password[0:]=password
    length_of_password=len(list_password)
    list_username[0:]=username
    count=1
    status=True
    if length_of_password==len(list_username):
        while status==True:
            uppercase=str(list_password[0]).upper()
            lowercase=str(list_password[0]).lower()
            if uppercase==list_username[0] or lowercase==list_username[0]:
                if count<length_of_password and list_password[count]==list_username[count]:
                    count+=1
                else:
                    status=False
            else:
                status=False
        if count==length_of_password:
            m="password contain username"
            
    return m
    
    
    

def passwordvalid(password):
    check = []
    digitValid = digitValidity(password)
    charValid = charValidity(password)
    lengthValid = lengthValidity(password)
    specialcharValid = specialcharValidity(password)
    char_not_included=checkchar(password)
    username_valid=usernameValid(password,username)
    check.extend([digitValid, charValid, lengthValid, specialcharValid,char_not_included,username_valid])
    n = len(check)
    count = 0
    for c in range(0, n):
        if check[c] == "":
            count += 1
    if count == n:
        print("password is accepted")
        return True
    else:
        print("password is not accepted because :")
        for c in range(0, n):
            if check[c] != "":
                print(check[c])
        return False


def emailformatValidity(email_address):
    m = ""
    count = 0
    checklist = ["@"]
    for c in email_address:
        if c in checklist:
            count += 1
            break
    if count != 1:
        m = "Incorrect Email Address format"
    return m


def emailValidity(email_address):
    check = []
    specialchar = emailformatValidity(email_address)
    check.extend([specialchar])
    n = len(check)
    count = 0
    for c in range(0, n):
        if check[c] == "":
            count += 1
    if count == n:
        print("email address accepted")
        return True
    else:
        print("email address is not accepted due to : ")
        for c in range(0, n):
            if check[c] != "":
                print(check[c])
        return False


def passwordcolumnsearch():
    for c in range(1, ws.max_column + 1):
        element = ws.cell(row=1, column=c)
        if element.value == "PASSWORD":
            k = c
            break
    return k

def usernamecolumnsearch():
    for c in range(1, ws.max_column + 1):
        element = ws.cell(row=1, column=c)
        if element.value == "USERNAME":
            q = c
            break
    return q

def emailaddresscolumnsearch():
    for c in range(1, ws.max_column + 1):
        element = ws.cell(row=1, column=c)
        if element.value == "EMAIL ADDRESS":
            r = c
            break
    return r

def fetch(c,k,q,r):
    element_username=ws.cell(row=c,column=q).value
    print("userid is : "+str(element_username))
    element_password=ws.cell(row=c,column=k).value
    print("password is : "+str(element_password))
    element_emailaddress=ws.cell(row=c,column=r).value
    print("email address is : "+str(element_emailaddress))

def migrate(sheetindex,g,c,q,k,r,status_sheet):
    if status_sheet=="changed":
        ws=wb[sheetindex[g]]
        username_copy=ws.cell(row=c,column=q).value
        password_copy=ws.cell(row=c,column=k).value
        emailaddress_copy=ws.cell(row=c,column=r).value
        ws.delete_rows(c)
        print("data is deleted from "+sheetindex[g]+ " sheet")
        if g==0:
            g=1
            ws=wb[sheetindex[g]]
            rowmax=ws.max_row
            username_paste=ws.cell(row=rowmax+1,column=q)
            username_paste.value=username_copy
            password_paste=ws.cell(row=rowmax+1,column=k)
            password_paste.value=password_copy
            emailaddress_paste=ws.cell(row=rowmax+1,column=r)
            emailaddress_paste.value=emailaddress_copy
            print("data is now added in "+sheetindex[g])
        else:
            g=0
            ws=wb[sheetindex[g]]
            rowmax=ws.max_row
            username_paste=ws.cell(row=rowmax+1,column=q)
            username_paste.value=username_copy
            password_paste=ws.cell(row=rowmax+1,column=k)
            password_paste.value=password_copy
            emailaddress_paste=ws.cell(row=rowmax+1,column=r)
            emailaddress_paste.value=emailaddress_copy
            print("data is now added in "+sheetindex[g])
        wb.save("test.xlsx")
    else:
        print("it exist in same sheet no need to migrate")



def existence(username,k,q,r,row1,g,status_sheet):
    password=""
    count=0
    for c in range(1, row1 + 1):
        val = ws.cell(row=c, column=q)
        boolean=False
        # checking if userId exists
        if username == val.value:
            count +=1
            boolean=True
            print("UserId exist in " + sheetindex[g] + " sheet")
            status="True"
            while (status!="False"):
                print("enter 1 to update details " + "\nenter 2 to fetch details "+"\nenter 3 to migrate user"+"\nenter 4 to exit")
                choice = input()
                if choice == "2":
                    fetch(c,k,q,r)
                if choice == "1":
                    password = input("enter your password\n")
                    m=passwordvalid(password)
                    element_update_password = ws.cell(row=c, column=k)
                    if m==True:
                        element_update_password.value = password
                    email_address = input("enter your email address\n")
                    j=emailValidity(email_address)
                    element_update_emailaddress = ws.cell(row=c, column=r)
                    if j==True:
                        element_update_emailaddress.value = email_address
                if choice=="3":
                    migrate(sheetindex,g,c,q,k,r,status_sheet)
                if choice=="4" or choice=="":
                    status="False"
                print("\n")
            break
    return count,boolean


def inputted(counter,k,q,r,row1):
    if counter != 1:
        password = input("enter your password\n")
        m=passwordvalid(password)
        if m==True:
            element_username = ws.cell(row=row1 + 1, column=q)
            element_username.value = username
            element_password = ws.cell(row=row1 + 1, column=k)
            element_password.value = password
            email_address = input("enter your email address\n")
            j=emailValidity(email_address)
            element_emailaddress = ws.cell(row=row1 + 1, column=r)
            if j==True:
                element_emailaddress.value = email_address

try:
    wb = op.load_workbook('test.xlsx')
    wb.active
    sheet=input("enter usertype \n")
    if sheet=="local" or sheet=="global":
        sheetindex=[]
        sheetindex.extend(wb.sheetnames)
        username = input("enter userId\n")
        user=usernamecheck(username)
        if user==True:
            g=0
            status_sheet="original"
            if sheet==sheetindex[0]:
                g=0
            else:
                g=1
            if sheet == "local" or sheet=="global":
                ws=wb[sheetindex[g]]
                row1=ws.max_row
                k=passwordcolumnsearch()
                q=usernamecolumnsearch()
                r=emailaddresscolumnsearch()
                counter,bul=existence(username,k,q,r,row1,g,status_sheet)
                if bul==False:
                    if g==0:
                        status_sheet="changed"
                        g=1
                        ws=wb[sheetindex[g]]
                        row1=ws.max_row
                        k=passwordcolumnsearch()
                        q=usernamecolumnsearch()
                        r=emailaddresscolumnsearch()
                        counter,bul=existence(username,k,q,r,row1,g,status_sheet)
                    else:
                        g=0
                        status_sheet="changed"
                        ws=wb[sheetindex[g]]
                        row1=ws.max_row
                        k=passwordcolumnsearch()
                        q=usernamecolumnsearch()
                        r=emailaddresscolumnsearch()
                        counter,bul=existence(username,k,q,r,row1,g,status_sheet)
            inputted(counter,k,q,r,row1)
            wb.save("test.xlsx")
        else:
            print(user)
    else:
        print("Invalid Usertype")
except Exception as e:
    print("Error in execution :" + str(e))