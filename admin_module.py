# This module retrieve data from admin database to send otp to verify the email 
import openpyxl as op
import random
import email_verification_module 


def admin_username(worksheet):
    for position in range(1, worksheet.max_column +1):
        element = worksheet.cell(row = 1, column = position)
        if element.value == "ADMIN USERNAME":
            break
    return position


def admin_password(worksheet):
    for position in range(1, worksheet.max_column+1):
        element = worksheet.cell(row=1, column = position)
        if element.value == "ADMIN PASSWORD":
            break
    return position


def email_address_search(worksheet, num , username_column, password_column):
    for position in range (1, worksheet.max_row+1):
        if position == num:
            element_emailaddress = worksheet.cell(row = position, column = username_column).value
            element_password = worksheet.cell(row = position, column = password_column).value
            break 
    return element_emailaddress, element_password


def main(recevier_email_address,username_userinput, password_userinput):
    workbook = op.load_workbook('confidential.xlsx')
    worksheet = workbook.active
    username_column = admin_username(worksheet)
    password_column = admin_password(worksheet)
    num = random.randint(2, worksheet.max_row)
    sender_email_address, sender_password = email_address_search(worksheet, num, username_column, password_column)
    otp,status = email_verification_module.otp_generator(recevier_email_address, sender_email_address, sender_password)
    if status == True:
        otp_inputted = input("enter the otp obtained\t")
        if otp_inputted == otp:
            email_verification_module.email_verified(recevier_email_address, sender_email_address, sender_password, username_userinput, password_userinput)
            return True



